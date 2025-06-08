"""
엑셀 파일 처리를 위한 유틸리티 클래스입니다.
"""

import pandas as pd
import logging
from typing import Dict, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelHandler:
    """엑셀 파일 처리를 위한 클래스"""
    
    def __init__(self):
        """ExcelHandler 초기화"""
        self.logger = logging.getLogger(__name__)
        self.styles = {
            'header': {
                'fill': PatternFill(start_color="00b0f0", 
                                  end_color="00b0f0", 
                                  fill_type="solid"),
                'font': Font(bold=True, color='FFFFFF')
            },
            'upper': {
                'fill': PatternFill(start_color="daeef3", 
                                  end_color="daeef3", 
                                  fill_type="solid")
            },
            'lower': {
                'fill': PatternFill(start_color="f2f2f2", 
                                  end_color="f2f2f2", 
                                  fill_type="solid")
            }
        }
    
    def _apply_styles(self, worksheet, style_type: str):
        """
        워크시트에 스타일을 적용합니다.
        
        Args:
            worksheet: 스타일을 적용할 워크시트
            style_type: 적용할 스타일 유형
        """
        if style_type == 'redundancy':
            # 헤더 스타일 적용
            for cell in worksheet[1]:
                cell.fill = self.styles['header']['fill']
                cell.font = self.styles['header']['font']
            
            # 데이터 스타일 적용
            for row in worksheet.iter_rows(min_row=2):
                style = self.styles['upper'] if row[1].value == 'Upper' else self.styles['lower']
                for cell in row:
                    cell.fill = style['fill']
        
        elif style_type == 'changes':
            # 헤더 스타일 적용
            for cell in worksheet[1]:
                cell.fill = self.styles['header']['fill']
                cell.font = self.styles['header']['font']
    
    def save_redundancy_analysis(self, 
                               df: pd.DataFrame, 
                               output_file: str):
        """
        중복 정책 분석 결과를 엑셀 파일로 저장합니다.
        
        Args:
            df: 저장할 데이터프레임
            output_file: 저장할 파일 경로
        """
        try:
            self.logger.info(f"중복 정책 분석 결과 저장 중: {output_file}")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                if 'vsys' in df.columns:
                    for vsys, vsys_df in df.groupby('vsys'):
                        sheet_name = f'Analysis_{vsys}'
                        vsys_df.to_excel(writer, 
                                       sheet_name=sheet_name, 
                                       index=False)
                        self._apply_styles(writer.sheets[sheet_name], 
                                         'redundancy')
                else:
                    df.to_excel(writer, 
                              sheet_name='Analysis', 
                              index=False)
                    self._apply_styles(writer.sheets['Analysis'], 
                                     'redundancy')
            
            self.logger.info(f"결과가 {output_file}에 저장되었습니다.")
            
        except Exception as e:
            self.logger.error(f"결과 저장 중 오류 발생: {e}")
            raise
    
    def save_change_analysis(self, 
                           results: Dict[str, pd.DataFrame], 
                           output_file: str):
        """
        변경사항 분석 결과를 엑셀 파일로 저장합니다.
        
        Args:
            results: 저장할 분석 결과 딕셔너리
            output_file: 저장할 파일 경로
        """
        try:
            self.logger.info(f"변경사항 분석 결과 저장 중: {output_file}")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 요약 정보 저장
                summary_data = {
                    'Category': ['추가된 정책', '제거된 정책', '변경된 정책'],
                    'Count': [len(results['added']), 
                            len(results['removed']), 
                            len(results['changed'])]
                }
                pd.DataFrame(summary_data).to_excel(writer, 
                                                  sheet_name='Summary', 
                                                  index=False)
                
                # 상세 정보 저장
                for sheet_name, df in results.items():
                    if not df.empty:
                        df.to_excel(writer, 
                                  sheet_name=sheet_name.capitalize(), 
                                  index=False)
                        self._apply_styles(writer.sheets[sheet_name.capitalize()], 
                                         'changes')
            
            self.logger.info(f"결과가 {output_file}에 저장되었습니다.")
            
        except Exception as e:
            self.logger.error(f"결과 저장 중 오류 발생: {e}")
            raise 